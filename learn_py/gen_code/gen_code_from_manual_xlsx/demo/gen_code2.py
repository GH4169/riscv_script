#!/usr/bin/env python3
# coding=utf-8

import re
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def expand_merged_cells(input_file, output_file):# {{{
    """
    展开 Excel 中所有合并单元格：将左上角的值复制填充到整个合并区域
    input_file: 输入的 Excel 文件路径
    output_file: 输出的展开后 Excel 文件路径
    """
    wb = load_workbook(input_file)
    ws = wb.active
    # ws = wb.worksheets[8]

    # 遍历所有合并单元格区域（先转换成 list 避免迭代修改报错）
    merged_ranges = list(ws.merged_cells.ranges)

    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        value = ws.cell(row=min_row, column=min_col).value

        # 先解除合并单元格
        ws.unmerge_cells(str(merged))

        # 填充整个原合并区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=value)

    wb.save(output_file)
    print(f"展开合并单元格完成，已保存到：{output_file}")# }}}

# 解析 WaveDrom 模板
def parse_wavedrom_adoc(file_path):# {{{
    templates = []
    with open(file_path, "r") as f:
        data = f.read()
    # 去掉 ```wavedrom 和 ```
    data = re.sub(r'```wavedrom', '', data)
    data = data.replace('```', '')

    # 按 {reg: 分段
    blocks = data.split("{reg:")
    for blk in blocks[1:]:
        blk = blk.strip()
        # 去掉开头和结尾的 []
        blk = blk.lstrip('[').rstrip(']}')

        # 提取每个字段行
        fields = []
        lines = blk.splitlines()
        for line in lines:
            line = line.strip().rstrip(",")
            if not line or line.startswith("//"):
                continue

            # 匹配 bits、name、type、attr
            m_bits = re.search(r'bits\s*:\s*(\d+)', line)
            m_name = re.search(r'name\s*:\s*(.*?)(,|})', line)
            m_type = re.search(r'type\s*:\s*(\d+)', line)
            m_attr = re.search(r'attr\s*:\s*(.*?)(}|$)', line)

            field = {}
            if m_bits:
                field['bits'] = int(m_bits.group(1))
            if m_name:
                # 去掉引号和空格
                name = m_name.group(1).strip()
                name = name.strip("'\]\"")
                name = re.sub(r"[\:\/\[\]\s]+", "_", name)
                # name = re.sub(r"_+", "_", name)
                field['name'] = name
            if m_type:
                field['type'] = int(m_type.group(1))
            if m_attr:
                attr = m_attr.group(1).strip()
                if attr.startswith("[") and attr.endswith("]"):
                    attr = attr[1:-1].strip()
                attr = attr.strip("'\"")
                # 如果是列表格式 ['OPIVI']
                if attr.startswith('[') and attr.endswith(']'):
                    attr = [a.strip("'\" ") for a in attr[1:-1].split(",")]
                field['attr'] = attr

            fields.append(field)

        templates.append({"reg": fields})
        # print({"reg": fields})
    # print(templates)
    return templates# }}}

# 根据 Excel 行选择模板
def select_template(templates, funct3_val):# {{{
    for t in templates:
        if 'attr' in t["reg"][0] and t["reg"][0].get("attr") == funct3_val:
            # print(t["reg"][0])
            return t["reg"]
    return None# }}}


def generate_code(template, row):# {{{
    """
    根据 WaveDrom template 和 Excel 行，生成 32bit 指令编码字符串
    - template: [{'bits': 7, 'name': '0x57'}, {'bits': 5, 'name': 'vd'}, ...]
    - row: Excel 一行，包含字段值
    """
    code_bits = []
    
    for f in template:
        name = f.get("name")
        width = f.get("bits", 0)
        
        # opcode / 常数字段，例如 0x57
        if name.startswith("0x"):
            val = bin(int(name, 16))[2:].zfill(width)
            code_bits.append(val)
            continue

        # 如果 name 本身是数字
        try:
            val_int = int(name, 0)  # 支持 10 进制或 0x 十六进制
            val_bin = bin(val_int)[2:].zfill(width)
            code_bits.append(val_bin)
            continue
        except:
            pass  # 非数字字段，继续处理 Excel 列匹配
        # 对应 Excel 列
        col_name = name.lower()  # 小写匹配
        val = row.get(col_name)
        
        if pd.isna(val) or str(val).strip() == "":
            # Excel 中没有值，用 ? 占位
            val_bin = "?" * width
        else:
            # Excel 本身就是二进制字符串，去掉空格并补齐
            val_bin = str(val).strip()
            # print(val_str)

        code_bits.append(val_bin)

    # 拼成 SystemVerilog 风格
    code_bits = code_bits[::-1]  # 反转列表
    return "32'b" + "_".join(code_bits)# }}}

def process_excel(template_file, input_excel, output_excel, output_fcov):# {{{
    templates = parse_wavedrom_adoc(template_file)
    df = pd.read_excel(input_excel, dtype=str)

    codes = []
    with open(output_fcov, "w") as f_fcov:
        for i, row in df.iterrows():
            funct3 = row.get("funct3")
            # print(funct3)
            tmpl = select_template(templates, funct3)
            if tmpl:
                code = generate_code(tmpl, row)
            else:
                raise ValueError(f"No matching template found for funct3={funct3}")
            
            codes.append(code)

            # 写入 fcov 文件
            assembly = row.get("assembly", "").strip()
            f_fcov.write(f"wildcard {assembly} = {{{code}}};\n")

    df["code"] = codes
    df.to_excel(output_excel, index=False)
    print("已生成：", output_excel)
    print("已生成：", output_fcov)# }}}


# 执行入口
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} op_format.adoc manual_v_inst.xlsx")
        sys.exit(1)
    os.makedirs("generated_v_inst", exist_ok=True)

    all_v_inst_filled_xlsx = os.path.join("generated_v_inst", "all_v_inst_filled.xlsx")
    expand_merged_cells(sys.argv[2], all_v_inst_filled_xlsx)

    output_all_v_inst_code_xlsx = os.path.join("generated_v_inst", "all_v_inst_code.xlsx")
    output_all_v_inst_fcov_sv = os.path.join("generated_v_inst", "all_v_inst_fcov.sv")
    process_excel(sys.argv[1], all_v_inst_filled_xlsx, output_all_v_inst_code_xlsx, output_all_v_inst_fcov_sv)
