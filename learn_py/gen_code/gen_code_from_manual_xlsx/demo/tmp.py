#!/usr/bin/env python3
# coding=utf-8
from openpyxl import load_workbook

def expand_merged_cells(input_file, output_file):
    """
    展开 Excel 中所有合并单元格：将左上角的值复制填充到整个合并区域
    input_file: 输入的 Excel 文件路径
    output_file: 输出的展开后 Excel 文件路径
    """
    wb = load_workbook(input_file)
    # ws = wb.active
    ws = wb.worksheets[8]

    # 遍历所有合并单元格区域（先转换成 list 避免迭代修改报错）
    merged_ranges = list(ws.merged_cells.ranges)

    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        value = ws.cell(row=min_row, column=min_col).value

        # 先解除合并单元格
        ws.unmerge_cells(str(merged))

        # 填充整个原合并区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=value)

    wb.save(output_file)
    print(f"展开合并单元格完成，已保存到：{output_file}")


# 示例调用
if __name__ == "__main__":
    expand_merged_cells("manual_v_inst.xlsx", "file_filled.xlsx")
